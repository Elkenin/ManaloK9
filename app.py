from flask import Flask, request, jsonify, render_template
import openpyxl
import uuid  # For generating unique IDs

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')  # Ensure index.html is in the "templates" folder

@app.route('/submit', methods=['POST'])
def submit_form():
    form_data = request.form
    uid = form_data.get('uniqueId') or str(uuid.uuid4())  # Generate UID if not provided
    data_to_save = [
        uid,  # Add UID to the first column
        form_data.get('status'),
        form_data.get('date'),
        form_data.get('name'),
        form_data.get('contact'),
        form_data.get('reservationDate'),
        form_data.get('in'),
        form_data.get('out'),
        form_data.get('day'),
        form_data.get('night'),
        form_data.get('receptionist'),
        form_data.get('adult'),
        form_data.get('children'),
        form_data.get('student'),
        form_data.get('videoke'),
        form_data.get('room'),
        form_data.get('roomNumber'),
        form_data.get('cottage'),
        form_data.get('noPax'),
        form_data.get('downpayment'),
        form_data.get('gcash'),
        form_data.get('fullPayment'),
        form_data.get('total')
    ]

    # Save to Excel
    file_name = "ManaloK9-Tracker.xlsx"
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if sheet.max_row == 1:  # Add headers if the sheet is empty
        headers = [
            "UID", "Status", "Date", "Name", "Contact", "Reservation Date",
            "Check-In", "Check-Out", "Day", "Night", "Receptionist",
            "Adult", "Children", "Student", "Videoke", "Room",
            "Room #", "Cottage", "No. Pax", "Downpayment", "GCash (Reference#)",
            "Full Payment", "Total"
        ]
        sheet.append(headers)

    # Update if UID exists, else add new row
    row_updated = False
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == uid:  # Match UID
            for i, value in enumerate(data_to_save):
                row[i].value = value
            row_updated = True
            break

    if not row_updated:
        sheet.append(data_to_save)

    workbook.save(file_name)

    return jsonify({"message": "Data saved successfully!", "data": get_excel_data(file_name)})

@app.route('/delete', methods=['POST'])
def delete_entry():
    uid_to_delete = request.form.get('uniqueId')
    if not uid_to_delete:
        return jsonify({"message": "Unique ID not provided!"}), 400

    file_name = "ManaloK9-Tracker.xlsx"
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        return jsonify({"message": "File not found!"}), 404

    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == uid_to_delete:  # Match UID
            sheet.delete_rows(row[0].row, 1)
            workbook.save(file_name)
            return jsonify({"message": "Entry deleted successfully!", "data": get_excel_data(file_name)})

    return jsonify({"message": "UID not found!"}), 404

def get_excel_data(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    return data

if __name__ == "__main__":
    app.run(debug=True)
